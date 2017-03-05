from os.path import exists
from openpyxl import load_workbook, Workbook, Worksheet
from openpyxl.styles import Alignment, Font
from typing import List, Any


class XlsxWriter(object):

    def __init__(self, sheet_path):
        self.sheet_path = sheet_path
        if exists(sheet_path):
            self.wb = load_workbook(self.sheet_path)
        else:
            self.wb = Workbook()

    def get_or_create_sheet(self, sheet_name: str, k_fold: int) -> Worksheet:
        try:
            ws = self.wb[sheet_name]
        except KeyError:
            ws = self.wb.create_sheet(sheet_name)
            ws.append(self.create_header(k_fold))
            self.format_header(ws)
            self.wb.save(self.sheet_path)

        return ws

    def format_header(self, ws: Worksheet)-> None:
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=11, bold=True)

    def format_results(self, ws: Worksheet) -> None:
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=11, bold=False)
            cell.number_format = '0.00000%'

    def create_header(self, k_fold: int) -> List[Any]:
        header = ["Algorithm"]
        for i in range(1, k_fold + 1):
            header.append("Fold " + str(i))
        header.append("Average")

        return header

    def append_results(self, sheet_name: str, algorithm: str, results: List[float], average: float) -> None:
        ws = self.get_or_create_sheet(sheet_name, len(results))
        row = [algorithm]
        row.extend(results)
        row.append(average)
        ws.append(row)
        self.format_results(ws)
        self.set_colum_width(ws)
        self.wb.save(self.sheet_path)

    def set_colum_width(self, ws: Worksheet) -> None:
        def as_text(value):
            if value is None:
                return ""
            return str(value)

        for column_cells in ws.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column].width = length + 6
